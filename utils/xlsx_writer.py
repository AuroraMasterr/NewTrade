from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os


def calculate_profit_pct(trade_log):
    profit_pct = trade_log["exit_price"] - trade_log["entry_price"]
    profit_pct = profit_pct * trade_log["leverage"] / trade_log["entry_price"]
    if trade_log["side"] == "开空":
        profit_pct = -profit_pct
    return profit_pct


def write_values(ws, row_idx, trade_log, current_cap, klines_figure):
    cell = ws.cell(row=row_idx, column=1, value=trade_log["pinbar_amplitude"])
    cell.number_format = "0.00%"
    ws.cell(row=row_idx, column=2, value=trade_log["side"])
    ws.cell(row=row_idx, column=3, value=str(trade_log["entry_dt"]))
    ws.cell(row=row_idx, column=4, value=trade_log["entry_price"])
    ws.cell(row=row_idx, column=5, value=trade_log["leverage"])
    ws.cell(row=row_idx, column=6, value=trade_log["tp_price"])
    ws.cell(row=row_idx, column=7, value=trade_log["sl_price"])
    ws.cell(row=row_idx, column=8, value=str(trade_log["exit_dt"]))
    ws.cell(row=row_idx, column=9, value=trade_log["exit_price"])
    cell = ws.cell(row=row_idx, column=10, value=calculate_profit_pct(trade_log))
    cell.number_format = "0.00%"
    cell = ws.cell(row=row_idx, column=11, value=current_cap)
    cell.number_format = "0.00"



def save_tradelog_to_xlsx(trade_log, klines_figure, xlsx_path, sheet_name, current_cap):
    headers = [
        "Pinbar振幅",
        "开仓方向",
        "开仓时间",
        "开仓价格",
        "杠杆倍数",
        "止盈线",
        "止损线",
        "平仓时间",
        "平仓价格",
        "收益率",
        "当前资金(初始100)",
        "持仓K线图",
    ]
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
    # 新表，补表头
    if ws.max_row == 1 and ws["A1"].value != headers[0]:
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=h)
    row_idx = ws.max_row + 1
    write_values(ws, row_idx, trade_log, current_cap, klines_figure)
    if klines_figure and os.path.exists(klines_figure):
        img = XLImage(klines_figure)
        img.width = 250
        img.height = 180
        img_anchor = f"{get_column_letter(12)}{row_idx}"
        ws.add_image(img, img_anchor)
        ws.row_dimensions[row_idx].height = 140.00  # 调整行高
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置内容对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = [
        12,
        12,
        22,
        12,
        12,
        12,
        12,
        22,
        12,
        12,
        12,
        22,
    ]
    for i, w in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(xlsx_path)


if __name__ == "__main__":
    trade_log = {
        "pinbar_amplitude": 0.01,
        "side": "开多",
        "entry_dt": "2026-05-06 16:28:40",
        "entry_price": 100,
        "leverage": 2,
        "tp_price": 115,
        "sl_price": 95,
        "exit_dt": "2026-05-06 17:00:00",
        "exit_price": 110,
        "start_bar": 1000,
        "exit_bar": 1200,
    }
    klines_figure = "pictures/chart_20260506_162840.png"
    save_tradelog_to_xlsx(
        trade_log,
        klines_figure,
        "test.xlsx",
        "test",
        100000,
    )
